#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
可视化fingeRNAt结果热图
展示每个residue的相互作用指纹
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # 使用非交互式后端
matplotlib.rcParams['font.family'] = 'Arial'
import matplotlib.pyplot as plt
import seaborn as sns
from matplotlib.colors import LinearSegmentedColormap
import os
import re
from pathlib import Path

def _parse_source_label(src: str):
    """Parse Source label like 'PZ38|8hb8LigandChainB.pdb|m1|lig:...' -> (source_stem, model_id)
    Returns (stem, model_id) or (None, None) if not parseable.
    """
    try:
        parts = str(src).split('|')
        if len(parts) < 2:
            return (None, None)
        source = parts[1]
        stem = Path(source).stem
        mid = None
        for p in parts:
            if p.startswith('m') and p[1:].isdigit():
                mid = int(p[1:])
                break
        return (stem, mid)
    except Exception:
        return (None, None)

def _build_index_map_from_rna(rna_pdb_path: Path):
    """Scan rna.pdb and build mapping (index, chain) -> (resseq, resname).
    Index increments when residue identity changes across the file.
    """
    if not rna_pdb_path.exists():
        return {}
    mapping = {}
    last = None
    idx = 0
    try:
        with rna_pdb_path.open('r') as fh:
            for line in fh:
                if not line.startswith(('ATOM',)):
                    continue
                resname = line[17:20].strip()
                chain = line[21].strip()
                resseq_str = line[22:26].strip()
                resseq = int(resseq_str) if resseq_str else None
                icode = line[26].strip()
                cur = (resname, chain, resseq, icode)
                if cur != last:
                    idx += 1
                    last = cur
                    mapping[(idx, chain)] = (resseq, resname)
    except Exception:
        return {}
    return mapping

def load_and_process_data(file_path):
    """加载并处理fingeRNAt数据"""
    
    print(f"读取数据文件: {file_path}")
    df = pd.read_csv(file_path, sep='\t')
    
    # 清理列名中的空格
    df.columns = df.columns.str.strip()
    # 规范列名：去掉前缀如 'rna.pdb#'，保留成 '1:A#HB' 的格式，便于解析
    def _norm_col(c: str) -> str:
        if isinstance(c, str) and '#'+c.split('#')[-1] in c:
            # 如果形如 'xxx#1:A#HB'，去掉最前面的片段直到第一个 '#'
            parts = c.split('#')
            if len(parts) >= 3 and parts[-2].count(':')==1:
                return parts[-2] + '#' + parts[-1]
        return c
    df = df.rename(columns={c: _norm_col(c) for c in df.columns})

    # 若没有 Source 列，使用 Ligand_name 作为 Source
    if 'Source' not in df.columns and 'Ligand_name' in df.columns:
        df['Source'] = df['Ligand_name']
    
    print(f"数据维度: {df.shape}")
    print(f"前几列: {list(df.columns[:5])}")
    
    # 清理数据：将非数值列转换为数值型
    basic_cols = ['Source', 'Ligand_name', 'is_solution']
    interaction_cols = [col for col in df.columns if col not in basic_cols]
    
    print(f"开始清理 {len(interaction_cols)} 个相互作用特征列...")
    
    for col in interaction_cols:
        # 将所有值转换为字符串，然后清理空格
        df[col] = df[col].astype(str).str.strip()
        # 将空字符串、'nan'、'NaN' 等转换为 0
        df[col] = df[col].replace(['', 'nan', 'NaN', 'None'], '0')
        # 转换为数值型
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    print("数据清理完成")
    return df

def extract_interaction_features(df):
    """提取相互作用特征列（兼容旧格式和 canonical 口袋索引）。"""
    
    # 找到相互作用特征列（排除基础信息列）
    basic_cols = ['Source', 'Ligand_name', 'is_solution']
    interaction_cols = [col for col in df.columns if col not in basic_cols]
    
    print(f"找到 {len(interaction_cols)} 个相互作用特征")
    
    # 解析特征名称，提取 residue 信息
    residue_info = []
    for col in interaction_cols:
        if not isinstance(col, str):
            continue
        # 1) 旧格式: "1:A#HB"
        m1 = re.match(r'(\d+):([A-Z])#(.+)', col)
        if m1:
            residue_num = int(m1.group(1))
            chain = m1.group(2)
            interaction_type = m1.group(3)
            residue_id = f"{residue_num}:{chain}"
        else:
            # 2) canonical 格式: "A:21|A:5#HB"
            m2 = re.match(r'^([^|]+)\|([^#]+)#(.+)$', col)
            if not m2:
                continue
            sol_idx = m2.group(1)
            pred_idx = m2.group(2)
            interaction_type = m2.group(3)
            # 将 canonical 口袋坐标整体作为 residue 维度标签
            residue_id = f"{sol_idx}|{pred_idx}"
            # residue_num/chain 对 canonical 无实际意义，用占位
            residue_num = 0
            chain = ''
        residue_info.append({
            'column': col,
            'residue_num': residue_num,
            'chain': chain,
            'interaction_type': interaction_type,
            'residue_id': residue_id,
        })
    
    residue_df = pd.DataFrame(residue_info)
    print(f"解析出 {len(residue_df)} 个residue特征")
    
    return df[interaction_cols], residue_df

def create_heatmap_by_residue(df, interaction_data, residue_info, output_dir='heatmaps', tsv_path: str=None):
    """按residue创建热图"""
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 获取唯一的residue和相互作用类型
    # canonical 模式下 residue_id 形如 "A:21|A:5"，直接按出现顺序排序；
    # 旧模式下 residue_id 形如 "10:A"，按链+index 排序。
    if not residue_info.empty:
        ids = list(residue_info['residue_id'].unique())
        if any('|' in rid for rid in ids):
            # 已经是 canonical 口袋标签，按字符串排序即可
            unique_residues = sorted(ids)
        else:
            chains = sorted(residue_info['chain'].unique())
            unique_residues = sorted(
                ids,
                key=lambda rid: (
                    next((i for i, c in enumerate(chains) if c == rid.split(':')[1]), 999),
                    int(rid.split(':')[0])
                )
            )
    else:
        unique_residues = []
    unique_interactions = sorted(residue_info['interaction_type'].unique())
    
    print(f"唯一residue数量: {len(unique_residues)}")
    print(f"相互作用类型: {unique_interactions}")
    
    # 创建residue × 相互作用类型的矩阵
    heatmap_data = []
    for _, row in df.iterrows():
        source = row['Source']
        residue_row = []
        
        for residue_id in unique_residues:
            for interaction_type in unique_interactions:
                # 查找对应的列
                matching_features = residue_info[
                    (residue_info['residue_id'] == residue_id) & 
                    (residue_info['interaction_type'] == interaction_type)
                ]
                
                if len(matching_features) > 0:
                    col_name = matching_features.iloc[0]['column']
                    value = row[col_name] if col_name in row else 0
                else:
                    value = 0
                
                residue_row.append(value)
        
        heatmap_data.append(residue_row)
    
    # 若可获取到参考来源的 rna.pdb，则构建 index→(resSeq,resName) 映射用于重标注
    label_map = {}
    if tsv_path is not None and len(df) > 0:
        # 取第一行 Source（我们在 summarize 中已将 solution 置顶）
        ref_src = str(df.iloc[0].get('Source', ''))
        stem, mid = _parse_source_label(ref_src)
        if stem and mid is not None:
            base = Path(tsv_path).resolve().parent  # puzzle 目录
            rna_pdb_path = base / stem / f"model_{int(mid):02d}" / 'rna.pdb'
            idx_map = _build_index_map_from_rna(rna_pdb_path)
            for rid in unique_residues:
                # canonical 口袋标签直接使用原始形式
                if '|' in rid:
                    label_map[rid] = rid
                    continue
                i_str, ch = rid.split(':')
                try:
                    i_val = int(i_str)
                except Exception:
                    i_val = None
                if i_val is not None:
                    res = idx_map.get((i_val, ch))
                    if res and res[0] is not None and res[1]:
                        label_map[rid] = f"{ch}:{res[0]}-{res[1]}"
                    else:
                        label_map[rid] = f"{ch}:{i_val}"
                else:
                    label_map[rid] = f"{ch}:{i_str}"

    # 创建列标签（应用重标注），对不同格式的 residue_id 做健壮处理
    def _pretty_residue_id(rid: str) -> str:
        if rid in label_map:
            return label_map[rid]
        parts = rid.split(':')
        # 旧格式 "10:A" → "A:10"
        if len(parts) == 2 and parts[0].isdigit():
            return f"{parts[1]}:{parts[0]}"
        # canonical 或其它复杂标签，直接返回原始
        return rid

    col_labels = []
    for residue_id in unique_residues:
        pretty = _pretty_residue_id(residue_id)
        for interaction_type in unique_interactions:
            col_labels.append(f"{pretty}_{interaction_type}")
    
    # 转换为DataFrame，并将纵轴标签从 Source 改为 "stem+modelid"（不含扩展名与 'm'）
    display_index = []
    for i, src in enumerate(df['Source']):
        stem, mid = _parse_source_label(str(src))
        lig = str(df.iloc[i]['Ligand_name']) if 'Ligand_name' in df.columns else ''
        if stem and mid is not None:
            label = f"{stem}{mid}"
        elif stem:
            label = stem
        else:
            label = str(src)
        if lig:
            label = f"{label} [{lig}]"
        display_index.append(label)

    heatmap_df = pd.DataFrame(
        heatmap_data,
        index=display_index,
        columns=col_labels
    )
    
    # 返回的 unique_residues 也替换为重标注后的顺序标签
    unique_residues_labeled = [_pretty_residue_id(rid) for rid in unique_residues]
    # Row colors: solution=green (#8cbe4f), original=blue (#025e90)
    row_flags = list(df['is_solution']) if 'is_solution' in df.columns else [0]*len(df)
    row_colors = ['#8cbe4f' if int(f)==1 else '#025e90' for f in row_flags]
    return heatmap_df, unique_residues_labeled, unique_interactions, row_colors


    


def _slugify(text: str) -> str:
    return re.sub(r'[^A-Za-z0-9._-]+', '_', str(text))[:120]


def plot_per_ligand_grids(heatmap_df, unique_residues, unique_interactions, output_dir):
    """Create per-ligand grid heatmaps (one image per Source)."""
    out_dir = Path(output_dir) / 'per_ligand'
    out_dir.mkdir(parents=True, exist_ok=True)

    n_interactions = len(unique_interactions)
    for source, row in heatmap_df.iterrows():
        # Build matrix residues x interactions
        values = []
        for residue_id in unique_residues:
            row_vals = []
            for interaction_type in unique_interactions:
                col = f"{residue_id}_{interaction_type}"
                row_vals.append(row.get(col, 0))
            values.append(row_vals)

        mat = pd.DataFrame(values, index=unique_residues, columns=unique_interactions)

        # Figure size scaling
        cell_w, cell_h = 0.4, 0.25
        fig_w = max(n_interactions * cell_w + 2, 6)
        fig_h = max(len(unique_residues) * cell_h + 2, 6)

        plt.figure(figsize=(fig_w, fig_h))
        colors = ['white', 'navy']
        cmap = LinearSegmentedColormap.from_list('binary', colors, N=2)
        sns.heatmap(mat, cmap=cmap, cbar=False, linewidths=0.2, linecolor='lightgray', vmin=0, vmax=1)
        plt.title(f'Ligand {source}', fontsize=12, pad=12)
        plt.xlabel('Interaction Type', fontsize=9)
        plt.ylabel('Residue', fontsize=9)
        plt.xticks(rotation=45, fontsize=7, ha='right')
        plt.yticks(fontsize=7)
        out_file = out_dir / f"ligand_{_slugify(source)}_grid.pdf"
        plt.tight_layout(pad=1.0)
        plt.savefig(out_file, dpi=300, bbox_inches='tight', facecolor='white')
        plt.close()


def export_pivot_summaries(heatmap_df, unique_residues, unique_interactions, out_xlsx):
    """Export pivot summaries to an Excel workbook:
    - per_ligand: counts by interaction type per ligand Source
    - per_residue: counts by interaction type per residue (aggregated across ligands)
    - totals: overall counts per interaction type and per residue
    """
    try:
        from openpyxl import Workbook
    except Exception:
        print('openpyxl not available, skipping pivot summaries')
        return

    wb = Workbook()
    # per_ligand
    ws1 = wb.active
    ws1.title = 'per_ligand'
    # Compute per-ligand counts
    lig_rows = []
    for source, row in heatmap_df.iterrows():
        counts = {}
        for it in unique_interactions:
            type_cols = [c for c in heatmap_df.columns if c.endswith(f'_{it}')]
            counts[it] = int(row[type_cols].sum())
        counts['Source'] = source
        lig_rows.append(counts)
    cols1 = ['Source'] + list(unique_interactions)
    ws1.append(cols1)
    for r in lig_rows:
        ws1.append([r.get(c, 0) for c in cols1])

    # per_residue
    ws2 = wb.create_sheet('per_residue')
    res_rows = []
    for res in unique_residues:
        counts = {}
        for it in unique_interactions:
            col = f"{res}_{it}"
            counts[it] = int(heatmap_df[col].sum()) if col in heatmap_df.columns else 0
        counts['Residue'] = res
        res_rows.append(counts)
    cols2 = ['Residue'] + list(unique_interactions)
    ws2.append(cols2)
    for r in res_rows:
        ws2.append([r.get(c, 0) for c in cols2])

    # totals
    ws3 = wb.create_sheet('totals')
    ws3.append(['metric', 'name', 'count'])
    for it in unique_interactions:
        type_cols = [c for c in heatmap_df.columns if c.endswith(f'_{it}')]
        ws3.append(['interaction_type', it, int(heatmap_df[type_cols].values.sum())])
    for res in unique_residues:
        cols = [c for c in heatmap_df.columns if c.startswith(f'{res}_')]
        ws3.append(['residue', res, int(heatmap_df[cols].values.sum()) if cols else 0])

    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f'Pivot summaries saved: {out_xlsx}')

def plot_comprehensive_heatmap(heatmap_df, unique_residues, unique_interactions, output_file, row_colors=None):
    """绘制综合热图，并在下方叠加每个残基位置的总指纹数量折线（对齐横轴）。"""
    n_cols = len(heatmap_df.columns)
    n_rows = len(heatmap_df.index)
    n_interactions = len(unique_interactions)
    n_residues = len(unique_residues)

    # 尺寸设置
    cell_w, cell_h = 0.3, 0.3
    fig_w = max(n_cols * cell_w + 5, 20)
    fig_h = max(n_rows * cell_h + 3, 10)
    fig = plt.figure(figsize=(fig_w, fig_h + 2))
    gs = fig.add_gridspec(2, 1, height_ratios=[5, 1], hspace=0.15)
    ax = fig.add_subplot(gs[0, 0])
    axp = fig.add_subplot(gs[1, 0], sharex=ax)

    # 绘制热图（双层叠加上色）
    if row_colors is not None and len(row_colors) == len(heatmap_df.index):
        saved_ylabels = list(heatmap_df.index)
        sol_mask_rows = [1 if (c.lower() == '#8cbe4f') else 0 for c in row_colors]
        import numpy as _np
        sol_df = heatmap_df.copy()
        sol_mask = _np.zeros_like(sol_df.values, dtype=bool)
        for i, flag in enumerate(sol_mask_rows):
            if flag == 0:
                sol_mask[i, :] = True
        sns.heatmap(sol_df, cmap=LinearSegmentedColormap.from_list('sol_bin', ['white', '#8cbe4f'], N=2),
                    cbar=False, xticklabels=True, yticklabels=saved_ylabels, linewidths=0.1, linecolor='lightgray',
                    vmin=0, vmax=1, mask=sol_mask, ax=ax)

        org_df = heatmap_df.copy()
        org_mask = _np.zeros_like(org_df.values, dtype=bool)
        for i, flag in enumerate(sol_mask_rows):
            if flag == 1:
                org_mask[i, :] = True
        sns.heatmap(org_df, cmap=LinearSegmentedColormap.from_list('org_bin', ['white', '#025e90'], N=2),
                    cbar=False, xticklabels=False, yticklabels=False, linewidths=0.1, linecolor='lightgray',
                    vmin=0, vmax=1, mask=org_mask, ax=ax)
        ax.set_yticks(_np.arange(len(saved_ylabels)) + 0.5)
        ax.set_yticklabels(saved_ylabels)
    else:
        sns.heatmap(heatmap_df, cmap=LinearSegmentedColormap.from_list('binary', ['white', 'navy'], N=2),
                    cbar=False, xticklabels=True, yticklabels=True, linewidths=0.1, linecolor='lightgray',
                    vmin=0, vmax=1, ax=ax)

    # 行标签着色
    if row_colors is not None:
        for i, lab in enumerate(ax.get_yticklabels()):
            if i < len(row_colors):
                lab.set_color(row_colors[i])

    ax.set_title('fingeRNAt Interaction Fingerprint Heatmap', fontsize=16, pad=12)
    ax.set_xlabel('')
    ax.set_ylabel('Structure Source', fontsize=8)
    ax.tick_params(axis='x', labelrotation=90, labelsize=6)
    ax.tick_params(axis='y', labelsize=6)

    # 残基分组边界
    for i in range(1, n_residues):
        ax.axvline(x=i * n_interactions, color='red', linewidth=1, alpha=0.6)

    # 计算每个残基位置的总指纹计数
    profile_vals = []
    x_centers = []
    for i, res in enumerate(unique_residues):
        cols = [f"{res}_{it}" for it in unique_interactions]
        total = int(heatmap_df[cols].values.sum())
        profile_vals.append(total)
        x_centers.append(i * n_interactions + n_interactions / 2)

    # 底部折线（山峰）
    axp.plot(x_centers, profile_vals, color='#444444', linewidth=1.5)
    axp.fill_between(x_centers, profile_vals, color='#999999', alpha=0.2)
    axp.set_ylabel('Total', fontsize=8)
    axp.set_xlabel('')
    axp.tick_params(axis='x', labelbottom=False)
    axp.grid(True, linestyle='--', linewidth=0.3, alpha=0.4)

    # 顶部标记相互作用类型（对所有残基重复标注）
    try:
        it_labels = []
        it_positions = []
        for j, col in enumerate(heatmap_df.columns):
            parts = str(col).rsplit('_', 1)
            it = parts[-1] if len(parts) == 2 else str(col)
            it_labels.append(it)
            it_positions.append(j + 0.5)
        ax_top = ax.twiny()
        ax_top.set_xlim(ax.get_xlim())
        ax_top.set_xticks(it_positions)
        ax_top.set_xticklabels(it_labels, rotation=90, fontsize=6, ha='center')
        ax_top.set_xlabel('Interaction Type (repeated per residue)', fontsize=8)
    except Exception:
        pass

    # 底部（折线图）标记 canonical 残基标签，明确 left/right 语义
    axp.set_xlabel('Residue (solution/reference | prediction; left = solution/reference, right = prediction)', fontsize=8)
    try:
        axp.set_xticks(x_centers)
        axp.set_xticklabels(unique_residues, rotation=90, fontsize=6)
    except Exception:
        pass

    plt.tight_layout()
    plt.savefig(output_file, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"综合热图（含底部指纹折线）已保存: {output_file}")
    plt.close(fig)

def plot_residue_specific_heatmaps(heatmap_df, unique_residues, unique_interactions, output_dir):
    """绘制每个residue的特定热图"""
    
    n_interactions = len(unique_interactions)
    
    for i, residue_id in enumerate(unique_residues[:10]):  # 只绘制前10个residue
        start_col = i * n_interactions
        end_col = (i + 1) * n_interactions
        
        residue_data = heatmap_df.iloc[:, start_col:end_col]
        residue_data.columns = unique_interactions
        
        # 计算固定色块大小的图片尺寸
        n_cols = len(residue_data.columns)
        n_rows = len(residue_data.index)
        
        cell_width = 0.5
        cell_height = 0.15
        
        fig_width = max(n_cols * cell_width + 2, 8)
        fig_height = max(n_rows * cell_height + 2, 6)
        
        plt.figure(figsize=(fig_width, fig_height))
        
        # 使用二值颜色映射
        colors = ['white', 'navy']
        cmap = LinearSegmentedColormap.from_list('binary', colors, N=2)
        
        sns.heatmap(
            residue_data,
            cmap=cmap,
            cbar=False,  # 不显示色卡
            annot=False,  # 不显示数值注释
            linewidths=0.2,
            linecolor='lightgray',
            vmin=0,
            vmax=1
        )
        
        # 设置固定字体大小
        plt.title(f'Residue {residue_id} Interaction Fingerprint', fontsize=12, pad=15)
        plt.xlabel('Interaction Type', fontsize=8)
        plt.ylabel('Structure Source', fontsize=8)
        
        # 强制6号字体
        plt.xticks(rotation=45, fontsize=6, ha='right')
        plt.yticks(fontsize=6, rotation=0)
        
        output_file = os.path.join(output_dir, f'residue_{residue_id.replace(":", "_")}_heatmap.pdf')
        plt.tight_layout(pad=1.0)
        plt.subplots_adjust(bottom=0.15, left=0.15)
        plt.savefig(output_file, dpi=300, bbox_inches='tight', facecolor='white')
        print(f"Residue {residue_id} 热图已保存: {output_file}")
        plt.close()

def _build_group_summary(heatmap_df, unique_interactions, df):
    """
    基于 df 的 Source / is_solution，将行按 solution / submitter 归组，
    对每个 group 计算每种相互作用的归一化总数（除以该 group 的模型数）。
    """
    if df is None or 'Source' not in df.columns:
        # Fallback: one row per structure（旧行为）
        interaction_summary = {}
        for interaction_type in unique_interactions:
            type_cols = [col for col in heatmap_df.columns if col.endswith(f'_{interaction_type}')]
            interaction_summary[interaction_type] = heatmap_df[type_cols].sum(axis=1)
        summary_df = pd.DataFrame(interaction_summary, index=heatmap_df.index)
        return summary_df

    groups = {}
    for i, row in df.iterrows():
        src = str(row.get('Source', ''))
        try:
            is_sol = int(row.get('is_solution', 0))
        except Exception:
            try:
                is_sol = 1 if float(row.get('is_solution', 0)) != 0.0 else 0
            except Exception:
                is_sol = 0
        stem, mid = _parse_source_label(src)
        stem = stem or src
        if is_sol == 1:
            gid = f"solution:{stem}"
            label = f"{stem} (solution)"
            gtype = 'solution'
        else:
            gid = f"submitter:{stem}"
            label = stem
            gtype = 'submitter'
        entry = groups.setdefault(gid, {'label': label, 'type': gtype, 'rows': [], 'models': set()})
        entry['rows'].append(i)
        if mid is not None:
            entry['models'].add((stem, mid))
        else:
            entry['models'].add((stem, i))

    # 先按出现顺序保留 solution，再按提交者名字排序
    ordered_ids = []
    seen = set()
    for gid, info in groups.items():
        if info['type'] == 'solution' and gid not in seen:
            ordered_ids.append(gid)
            seen.add(gid)
    submitters = sorted(
        [gid for gid, info in groups.items() if info['type'] == 'submitter'],
        key=lambda k: groups[k]['label']
    )
    for gid in submitters:
        if gid not in seen:
            ordered_ids.append(gid)
            seen.add(gid)

    summary_rows = []
    row_labels = []
    for gid in ordered_ids:
        info = groups[gid]
        rows_idx = info['rows']
        model_count = max(1, len(info['models']))
        vals = {}
        for interaction_type in unique_interactions:
            type_cols = [col for col in heatmap_df.columns if col.endswith(f'_{interaction_type}')]
            if not type_cols:
                vals[interaction_type] = 0.0
                continue
            sub = heatmap_df.iloc[rows_idx][type_cols]
            total = float(sub.values.sum())
            vals[interaction_type] = total / model_count
        summary_rows.append(vals)
        row_labels.append(info['label'])

    summary_df = pd.DataFrame(summary_rows, index=row_labels)
    return summary_df


def plot_interaction_summary(heatmap_df, unique_interactions, output_file, df=None):
    """绘制按 solution / submitter 聚合并归一化的相互作用类型汇总图。"""
    summary_df = _build_group_summary(heatmap_df, unique_interactions, df)

    n_interactions_types = len(summary_df.columns)
    n_groups = len(summary_df.index)

    fig_width = max(n_interactions_types * 0.8 + 2, 8)
    fig_height = max(n_groups * 0.6 + 2, 6)

    plt.figure(figsize=(fig_width, fig_height))

    ax = sns.heatmap(
        summary_df,
        cmap='Reds',
        cbar=True,
        xticklabels=True,
        yticklabels=True,
        linewidths=0.1,
        linecolor='white'
    )

    plt.title('Interaction Type Summary by Solution / Submitter', fontsize=14, pad=16)
    plt.xlabel('Interaction Type', fontsize=10)
    plt.ylabel('Group (solution / submitter)', fontsize=10)
    plt.xticks(rotation=45, fontsize=8, ha='right')
    plt.yticks(fontsize=8)

    plt.tight_layout(pad=1.0)
    plt.savefig(output_file, dpi=300, bbox_inches='tight', facecolor='white')
    print(f"相互作用汇总热图已保存: {output_file}")
    plt.close()


def plot_by_interaction_heatmaps(heatmap_df, unique_residues, unique_interactions, output_dir, row_colors=None):
    """For each interaction type, plot a heatmap with
    - X axis: residue ids (e.g., 1:A, 2:A, ...)
    - Y axis: structure source (Source)
    - Cell value: 0/1 presence of that interaction at that residue for that structure
    """
    out_dir = Path(output_dir) / 'by_interaction'
    out_dir.mkdir(parents=True, exist_ok=True)

    # Ensure source order is stable
    sources = list(heatmap_df.index)
    residues = list(unique_residues)

    for it in unique_interactions:
        # Build matrix with columns=residues, rows=sources
        cols = [f"{r}_{it}" for r in residues]
        # Some residues may be missing for this interaction; fill zeros
        sub = heatmap_df.reindex(columns=cols, fill_value=0)
        # Figure size scaling
        n_cols = len(residues)
        n_rows = len(sources)
        cell_w = 0.35
        cell_h = 0.05
        fig_w = max(n_cols * cell_w + 2, 10)
        fig_h = max(n_rows * cell_h + 2, 8)

        plt.figure(figsize=(fig_w, fig_h))
        colors = ['white', 'navy']
        cmap = LinearSegmentedColormap.from_list('binary', colors, N=2)
        if row_colors is not None and len(row_colors) == len(sources):
            sol_mask_rows = [1 if (c.lower() == '#8cbe4f') else 0 for c in row_colors]
            # solution 层
            sol_df = sub.copy()
            import numpy as _np
            sol_mask = _np.zeros_like(sol_df.values, dtype=bool)
            for i, flag in enumerate(sol_mask_rows):
                if flag == 0:
                    sol_mask[i, :] = True
            ax = sns.heatmap(
                sol_df,
                cmap=LinearSegmentedColormap.from_list('sol_bin', ['white', '#8cbe4f'], N=2),
                cbar=False,
                xticklabels=residues,
                yticklabels=sources,
                vmin=0,
                vmax=1,
                linewidths=0.1,
                linecolor='lightgray',
                mask=sol_mask
            )
            # original 层
            org_df = sub.copy()
            org_mask = _np.zeros_like(org_df.values, dtype=bool)
            for i, flag in enumerate(sol_mask_rows):
                if flag == 1:
                    org_mask[i, :] = True
            sns.heatmap(
                org_df,
                cmap=LinearSegmentedColormap.from_list('org_bin', ['white', '#025e90'], N=2),
                cbar=False,
                xticklabels=residues,
                yticklabels=False,
                vmin=0,
                vmax=1,
                linewidths=0.1,
                linecolor='lightgray',
                mask=org_mask,
                ax=ax
            )
            # 恢复纵轴刻度与标签
            ax.set_yticks(_np.arange(len(sources)) + 0.5)
            ax.set_yticklabels(sources)
        else:
            ax = sns.heatmap(
                sub,
                cmap=cmap,
                cbar=False,
                xticklabels=residues,
                yticklabels=sources,
                vmin=0,
                vmax=1,
                linewidths=0.1,
                linecolor='lightgray'
            )
        if row_colors is not None:
            ytick = ax.get_yticklabels()
            for i, lab in enumerate(ytick):
                if i < len(row_colors):
                    lab.set_color(row_colors[i])
        plt.title(f'Interaction {it}', fontsize=14, pad=14)
        plt.xlabel('Residue', fontsize=10)
        plt.ylabel('Structure Source', fontsize=10)
        plt.xticks(rotation=45, fontsize=7, ha='right')
        plt.yticks(fontsize=7)
        out_file = out_dir / f"interaction_{_slugify(it)}_heatmap.pdf"
        plt.tight_layout(pad=1.0)
        plt.savefig(out_file, dpi=300, bbox_inches='tight', facecolor='white')
        print(f'By-interaction heatmap saved: {out_file}')
        plt.close()

def create_statistical_summary(df, heatmap_df, output_file):
    """创建统计汇总"""
    
    stats = {
        'total_structures': len(df),
        'total_features': heatmap_df.shape[1],
        'mean_interaction_per_structure': heatmap_df.sum(axis=1).mean(),
        'std_interaction_per_structure': heatmap_df.sum(axis=1).std(),
        'max_interaction_value': heatmap_df.max().max(),
        'min_interaction_value': heatmap_df.min().min()
    }
    
    # 按来源统计
    source_stats = df.groupby('Source').size().to_dict()
    
    with open(output_file, 'w') as f:
        f.write("fingeRNAt结果统计汇总\n")
        f.write("=" * 50 + "\n\n")
        
        f.write("总体统计:\n")
        for key, value in stats.items():
            f.write(f"  {key}: {value:.3f}\n")
        
        f.write("\n按来源统计:\n")
        for source, count in source_stats.items():
            f.write(f"  {source}: {count}\n")
        
        f.write(f"\n最活跃的10个特征:\n")
        feature_sums = heatmap_df.sum(axis=0).sort_values(ascending=False)
        for i, (feature, value) in enumerate(feature_sums.head(10).items()):
            f.write(f"  {i+1}. {feature}: {value:.3f}\n")
    
    print(f"统计汇总已保存: {output_file}")

def main():
    """主函数"""
    
    print("=" * 60)
    print("fingeRNAt结果热图可视化")
    print("=" * 60)
    
    # 数据文件路径
    data_file = 'fingernaat_summary_results.tsv'
    
    if not os.path.exists(data_file):
        print(f"错误: 数据文件不存在: {data_file}")
        return
    
    # 创建输出目录
    output_dir = 'heatmaps'
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 1. 加载数据
    df = load_and_process_data(data_file)
    
    # 2. 提取相互作用特征
    interaction_data, residue_info = extract_interaction_features(df)
    
    # 3. 创建热图数据
    heatmap_df, unique_residues, unique_interactions = create_heatmap_by_residue(
        df, interaction_data, residue_info, output_dir
    )
    
    # 4. 绘制综合热图
    plot_comprehensive_heatmap(
        heatmap_df,
        unique_residues,
        unique_interactions,
        os.path.join(output_dir, 'comprehensive_heatmap.pdf')
    )
    
    # 5. 绘制相互作用类型汇总图
    plot_interaction_summary(
        heatmap_df,
        unique_interactions,
        os.path.join(output_dir, 'interaction_summary_heatmap.pdf'),
        df=df,
    )
    
    # 6. 绘制特定residue热图（前10个）
    plot_residue_specific_heatmaps(
        heatmap_df,
        unique_residues,
        unique_interactions,
        output_dir
    )
    
    # 7. 创建统计汇总
    create_statistical_summary(
        df,
        heatmap_df,
        os.path.join(output_dir, 'statistics_summary.txt')
    )
    
    print("\n" + "=" * 60)
    print("热图可视化完成！")
    print(f"所有输出文件保存在: {output_dir}")
    print("=" * 60)

if __name__ == "__main__":
    main() 
