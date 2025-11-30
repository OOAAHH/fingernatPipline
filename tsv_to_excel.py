# -*- coding: utf-8 -*-
"""
将 TSV 文件转换为 Excel：
- 可选地删除指定列；
- 在指纹列中，将字符串中的每个字符 "1" 标记为红色，"0" 保持黑色。

使用方法（在当前目录下执行）：
    python3 tsv_to_excel.py

如需调整输入 / 输出文件名或列配置，请修改 MAIN_CONFIG。
"""

import os
import re

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock, InlineFont
from openpyxl.styles import Alignment


MAIN_CONFIG = {
    # 输入 TSV 文件路径
    "input_tsv": "PZ21_fingernaat_summary_results_binding_pocket.tsv",
    # 输出 Excel 文件路径
    "output_xlsx": "PZ21_fingernaat_summary_results_binding_pocket.xlsx",
    # 是否在表格中展示 DataFrame 的 index
    "show_index": True,
    # 要移除的列名列表
    # 默认：移除 Ligand_name，保留 Source 与 is_solution
    "columns_to_remove": ["Ligand_name"],
    # 指纹列的起始索引（从0开始，不包括 index 列，移除列后需要调整）
    # 典型 binding_pocket TSV 结构：Source, Ligand_name, is_solution, <res1>, <res2>, ...
    # 移除 Ligand_name 后，列顺序为：Source, is_solution, <res1>, <res2>, ...
    # 因此指纹列从索引 2 开始。
    "fingerprint_start_col": 2,
}


def read_tsv_to_dataframe(tsv_path, columns_to_remove=None):
    """
    读取 TSV 文件为 pandas DataFrame，并移除指定的列。
    所有内容都读为字符串，便于后续处理。
    """
    if not os.path.exists(tsv_path):
        raise FileNotFoundError(f"TSV 文件不存在: {tsv_path}")

    df = pd.read_csv(tsv_path, sep="\t", dtype=str)

    if columns_to_remove:
        existing_cols = [col for col in columns_to_remove if col in df.columns]
        if existing_cols:
            df = df.drop(columns=existing_cols)
            print(f"已移除列: {', '.join(existing_cols)}")

    return df


def is_fingerprint(text):
    """
    判断文本是否为指纹格式（只包含 0 和 1 的较长字符串）。
    """
    if text is None:
        return False
    if not isinstance(text, str):
        text = str(text)
    text = text.strip()
    return bool(re.match(r"^[01]{10,}$", text))


def build_fingerprint_rich_text(text):
    """
    将指纹字符串转换为单元格富文本：
    - 连续的 '1' 用红色字体；
    - 连续的 '0' 用黑色字体；
    这样在一个单元格内就可以实现字符级别的着色。
    """
    if text is None:
        return ""
    if not isinstance(text, str):
        text = str(text)
    if not text:
        return ""

    # 按连续字符分段，减少 TextBlock 数量
    segments = []
    current_char = text[0]
    current_segment = [current_char]
    for ch in text[1:]:
        if ch == current_char:
            current_segment.append(ch)
        else:
            segments.append(("".join(current_segment), current_char))
            current_char = ch
            current_segment = [ch]
    segments.append(("".join(current_segment), current_char))

    blocks = []
    for seg_text, ch in segments:
        # 使用 InlineFont（而不是 styles.Font），这是 rich_text.TextBlock 所要求的类型
        if ch == "1":
            font = InlineFont(color="FFFF0000")  # 红色
        else:
            font = InlineFont(color="FF000000")  # 黑色
        blocks.append(TextBlock(font=font, text=seg_text))

    return CellRichText(*blocks)


def dataframe_to_excel(
    df,
    output_xlsx,
    show_index=True,
    fingerprint_start_col=1,
):
    """
    将 DataFrame 写入 Excel，并在指纹列中对字符串中的 '1' 着红色。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 统一的 index 解释行文本（用于宽度基准）
    index_expl = "This row aligns solution and prediction residues in the canonical pocket index"

    # 在首行之上写入指纹说明
    legend_lines = [
        "Fingerprint bit legend (per residue column): "
        "1: HB, 2: HAL, 3: CA, 4: Pi_Cation, 5: Pi_Anion, "
        "6: Pi_Stacking, 7: Mg_mediated, 8: K_mediated, "
        "9: Na_mediated, 10: Other_mediated, 11: Water_mediated, 12: Lipophilic",
    ]

    # 写 legend 到第一行，并将前 3 个单元格合并居中
    legend_row = 1
    ws.cell(row=legend_row, column=1, value=legend_lines[0])
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=3)
    ws.cell(row=legend_row, column=1).alignment = Alignment(
        wrap_text=True, horizontal="center", vertical="center"
    )

    header_row = legend_row + 1

    # 写表头
    col_offset = 0
    if show_index:
        ws.cell(row=header_row, column=1, value="Index")
        col_offset = 1

    for col_idx, col_name in enumerate(df.columns):
        ws.cell(row=header_row, column=col_offset + col_idx + 1, value=str(col_name))

    # 写数据行
    first_data_row = header_row + 1
    expl_row = None
    for row_num, (idx, row) in enumerate(df.iterrows(), start=first_data_row):
        # index 列
        if show_index:
            index_value = idx
            # 若这是第一行数据且包含 solution=...; prediction=... 注释，则在最左侧给出解释
            if row_num == first_data_row:
                try:
                    row_values = list(row.values)
                except Exception:
                    row_values = []
                if any(isinstance(v, str) and "solution=" in v for v in row_values):
                    index_value = index_expl
                    expl_row = row_num
            ws.cell(row=row_num, column=1, value=index_value)

        # 数据列
        for col_idx, value in enumerate(row):
            excel_col = col_offset + col_idx + 1
            text = "" if value is None else str(value)

            # 判断是否为指纹列
            if col_idx >= fingerprint_start_col and is_fingerprint(text):
                rich_value = build_fingerprint_rich_text(text)
                ws.cell(row=row_num, column=excel_col).value = rich_value
            else:
                ws.cell(row=row_num, column=excel_col, value=text)

    # 简单设置一下列宽（根据最大字符串长度粗略估计）
    from openpyxl.utils import get_column_letter

    col_widths = {}
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # 允许为特定列使用固定列宽
        if col_idx == 1:
            width = 5.08
        elif col_idx == 2:
            width = 77.15
        elif col_idx == 3:
            width = 10.77
        else:
            max_len = 0
            for cell in ws[col_letter]:
                cell_value = cell.value
                if cell_value is None:
                    continue
                if isinstance(cell_value, CellRichText):
                    # 富文本的原始字符串
                    raw_text = "".join(block.text for block in cell_value)
                else:
                    raw_text = str(cell_value)
                max_len = max(max_len, len(raw_text))
            width = max(10, max_len * 1.2)
        ws.column_dimensions[col_letter].width = width
        col_widths[col_idx] = width

    # 为首列开启自动换行，使 legend 在单元格内换行展示
    col_letter = get_column_letter(1)
    for cell in ws[col_letter]:
        cell.alignment = Alignment(wrap_text=True)

    # 如果存在 index 解释行，则合并该行前 3 个单元格并居中显示，不换行
    if expl_row is not None:
        ws.merge_cells(start_row=expl_row, start_column=1, end_row=expl_row, end_column=3)
        cell = ws.cell(row=expl_row, column=1)
        cell.alignment = Alignment(wrap_text=False, horizontal="center", vertical="center")

    wb.save(output_xlsx)
    print(f"✓ Excel 已成功生成: {output_xlsx}")
    print("  - 指纹列中每个 '1' 已标记为红色，'0' 为黑色")


def make_excel_from_tsv(
    input_tsv,
    output_xlsx,
    show_index=True,
    fingerprint_start_col=1,
    columns_to_remove=None,
):
    """
    一步完成：从 TSV 读取数据并生成带颜色指纹的 Excel。
    """
    df = read_tsv_to_dataframe(input_tsv, columns_to_remove=columns_to_remove)
    dataframe_to_excel(
        df=df,
        output_xlsx=output_xlsx,
        show_index=show_index,
        fingerprint_start_col=fingerprint_start_col,
    )


def main():
    cfg = MAIN_CONFIG
    make_excel_from_tsv(
        input_tsv=cfg["input_tsv"],
        output_xlsx=cfg["output_xlsx"],
        show_index=cfg["show_index"],
        fingerprint_start_col=cfg["fingerprint_start_col"],
        columns_to_remove=cfg["columns_to_remove"],
    )


if __name__ == "__main__":
    main()
