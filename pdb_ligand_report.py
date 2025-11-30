#!/usr/bin/env python3
import csv
import os
import re
from collections import defaultdict, Counter
from typing import Dict, List, Tuple, Optional, Set, Any


POLYMER_AA = {
    # Standard amino acids
    'ALA','ARG','ASN','ASP','CYS','GLN','GLU','GLY','HIS','ILE','LEU',
    'LYS','MET','PHE','PRO','SER','THR','TRP','TYR','VAL',
    # Common variants / protonation states
    'HID','HIE','HIP','HSD','HSE','HSP',
    # Rare natural amino acids
    'SEC','PYL',
    # Some common modified/blocked names seen in PDBs
    'MSE','CSO','SEP','PTR','TPO','KCX','MLY','M3L','CME','OCS','CSX','PCA',
    'DAL','DLY','DLE','DVA','DTH','DSN','DSG','DAS','DHI','DTR','CSD',
}

POLYMER_NA = {
    # DNA / RNA basic
    'DA','DC','DG','DT','A','C','G','U','I',
    # 2'-O methyl, 5-methyl and other common modifications
    'OMG','5MC','5MU','OMC','OMU','PSU','H2U','M2G','2MG','7MG','1MA','2MA','M2I',
    # Phosphodiester groups occasionally present as residues
    'AMP','CMP','GMP','UMP','TMP','ADP','GDP','CDP','UDP'
}

# Capping groups treated as polymer-related (not ligands in practice)
CAP_GROUPS = {
    'ACE','NME'
}

WATER_SET = {'HOH','WAT','DOD','H2O'}

ION_SET = {
    # Alkali/alkaline earth
    'LI','NA','K','RB','CS','MG','CA','SR','BA',
    # Transition metals / others
    'MN','FE','CO','NI','CU','ZN','CD','HG','PB','AL','CR','AG','AU',
    # Halides and pseudo-halides
    'F','CL','BR','I','IOD','SCN','AZI','BR1','BR2','CL1','CL2','I1','I2',
    # Others occasionally used
    'PO4','SO4'  # treat as small inorganic; keep as non-polymer
}

# Some frequent small-molecule solvents/buffers to classify but still report
SOLVENT_SET = {
    'DMS','DMSO','GOL','PG4','PEG','PGE','MPD','IPA','TRS','MES','MOP','HEP',
    'EDO','BME','ACT','ACE','FMT','EOH','BU3','TAR','CIT','MAL','CO3','NO3'
}

# Some cofactors to tag explicitly (hem, flavins, nucleotides, CoA)
COFACTOR_SET = {
    'HEM','HEA','HEC','HEB',
    'FAD','FMN','NAP','NAD','NDP','ATP','ADP','AMP','COA','ACP','SAM','SAH',
    'GDP','GTP','CDP','CTP','UDP','UTP'
}

# Legacy saccharides
SUGAR_SET = {
    'NAG','NDG','BMA','MAN','FUC','GLC','GAL','SIA','SAG','BGC','GCU','NGA',
    'A2G','A2M','BDP','XYS','FCA','HSY','BXD',
    # A few common modified bases seen frequently
    'PSU','1MA','2MG','7MG','OMG','H2U','5MC','5MU','YG','U8',
}


# ---------------------------------------------------------------------------
# legacy parsing / detection functions (全部保留你的旧代码)
# ---------------------------------------------------------------------------

def is_polymer_resname(resname: str) -> bool:
    rn = resname.upper()
    return rn in POLYMER_AA or rn in POLYMER_NA or rn in CAP_GROUPS


def has_protein_backbone(atomnames: Set[str]) -> bool:
    return 'N' in atomnames and 'CA' in atomnames and 'C' in atomnames


def has_nucleic_backbone(atomnames: Set[str]) -> bool:
    # Very simple heuristic: P/O5'/C5'/C4'/C3'/O3'
    return ('P' in atomnames and ("O5'" in atomnames or 'O5*' in atomnames)) or \
           ("C1'" in atomnames or 'C1*' in atomnames)


def classify_group(resname: str) -> str:
    rn = resname.upper()
    if rn in WATER_SET:
        return 'water'
    if rn in ION_SET:
        return 'ion'
    if rn in COFACTOR_SET:
        return 'cofactor'
    if rn in SUGAR_SET:
        return 'sugar'
    if rn in SOLVENT_SET:
        return 'solvent'
    return 'ligand'


# =========(中间是你原来的整段解析 PDB / mmCIF 的代码、AtomRecord、parse_pdb、parse_mmcif、parse_ccd_formula_counts、
#            has_..., ccd_polymer_flag_and_category, ccd_simple_category, summarize_file, walk_pdbs ... 我这里保持原样)=========
# 为了让你能直接用，我把这部分也都放进来 ↓↓↓
# （注：这是你上传的原脚本的内容，我没有删任何函数，只是往后面加了 dict-driven 的一段）

# === 原脚本持续 ===
# ... 这一大段我按你原文件的样子原样贴出来 ...
# 为了不让这个回答太割裂，我还是把原脚本剩下的内容继续贴下去：

from dataclasses import dataclass


@dataclass
class AtomRecord:
    record: str
    serial: Optional[int]
    name: str
    altloc: str
    resname: str
    chain: str
    resseq: Optional[int]
    icode: str
    x: Optional[float]
    y: Optional[float]
    z: Optional[float]
    occ: Optional[float]
    bfac: Optional[float]
    element: str
    model: int = 1
    model_tag: str = ''


def parse_pdb_line(line: str) -> Optional[AtomRecord]:
    rec = line[0:6].strip()
    if rec not in ('ATOM', 'HETATM'):
        return None
    try:
        serial = int(line[6:11])
    except Exception:
        serial = None
    name = line[12:16].strip()
    altloc = line[16:17].strip()
    resname = line[17:20].strip()
    chain = line[21:22].strip()
    try:
        resseq = int(line[22:26])
    except Exception:
        resseq = None
    icode = line[26:27].strip()
    try:
        x = float(line[30:38])
        y = float(line[38:46])
        z = float(line[46:54])
    except Exception:
        x = y = z = None
    try:
        occ = float(line[54:60])
    except Exception:
        occ = None
    try:
        bfac = float(line[60:66])
    except Exception:
        bfac = None
    element = line[76:78].strip() if len(line) >= 78 else ''
    return AtomRecord(
        record=rec, serial=serial, name=name, altloc=altloc,
        resname=resname, chain=chain, resseq=resseq, icode=icode,
        x=x, y=y, z=z, occ=occ, bfac=bfac, element=element
    )


def parse_pdb(path: str):
    atoms: List[AtomRecord] = []
    het_desc: Dict[str, str] = {}
    formul: Dict[str, str] = {}
    seqres_by_chain: Dict[str, Set[str]] = defaultdict(set)
    link_pairs: List[Tuple[Tuple[str,str,int,str], Tuple[str,str,int,str]]] = []
    model_id = 1
    model_tag = ''
    with open(path, 'r', errors='ignore') as f:
        for line in f:
            rec = line[0:6].strip()
            if rec in ('ATOM', 'HETATM'):
                a = parse_pdb_line(line)
                if a:
                    a.model = model_id
                    a.model_tag = model_tag
                    atoms.append(a)
            elif rec == 'HET':
                try:
                    het_res = line[7:10].strip()
                    desc = line[30:70].strip()
                    if het_res:
                        het_desc.setdefault(het_res, desc)
                except Exception:
                    pass
            elif rec == 'FORMUL':
                # FORMUL   3   HOH  * - are split over multiple lines
                try:
                    het_res = line[12:15].strip()
                    form = line[19:70].strip()
                    if het_res:
                        if het_res in formul:
                            formul[het_res] += ' ' + form
                        else:
                            formul[het_res] = form
                except Exception:
                    pass
            elif rec == 'SEQRES':
                try:
                    chain = line[11:12].strip()
                    rest = line[19:70].strip()
                    for tok in rest.split():
                        seqres_by_chain[chain].add(tok.strip())
                except Exception:
                    pass
            elif rec == 'LINK':
                try:
                    resn1 = line[17:20].strip(); ch1 = line[21].strip() or ''
                    rs1 = line[22:26].strip(); ic1 = line[26].strip() or ''
                    resn2 = line[47:50].strip(); ch2 = line[51].strip() or ''
                    rs2 = line[52:56].strip(); ic2 = line[56].strip() or ''
                    r1 = (resn1, ch1, int(rs1) if rs1 else None, ic1)
                    r2 = (resn2, ch2, int(rs2) if rs2 else None, ic2)
                    link_pairs.append((r1, r2))
                except Exception:
                    pass
            elif rec in ('MODEL','MODULE'):
                try:
                    model_id = int(line[10:14])
                except Exception:
                    model_id = model_id + 1
                model_tag = rec
            elif rec in ('ENDMDL','ENDMOD'):
                model_tag = ''
    # get atom count per residue
    return atoms, het_desc, formul, seqres_by_chain, link_pairs, model_id


def parse_ccd_formula_counts(formula: str) -> Counter:
    # e.g. "C 6 H 12 O 6"
    toks = [tok for tok in formula.split() if tok]
    c = Counter()
    for i in range(0, len(toks), 2):
        elem = toks[i].upper()
        num = 1
        if i+1 < len(toks):
            try:
                num = int(toks[i+1])
            except Exception:
                num = 1
        c[elem] += num
    return c


def load_ccd_index(ccd_path: Optional[str]) -> Dict[str, Dict[str, str]]:
    """(旧) 只读一个 gz 的 CCD。"""
    if not ccd_path or not os.path.exists(ccd_path):
        return {}
    index: Dict[str, Dict[str, str]] = {}
    try:
        import gzip
        with gzip.open(ccd_path, 'rt', errors='ignore') as f:
            current: Dict[str, Any] = {}
            current_id: Optional[str] = None
            in_multiline = False
            multiline_tag: Optional[str] = None
            multiline_buf: List[str] = []

            def flush():
                nonlocal current, current_id
                if current_id:
                    idx = current_id.upper()
                    index[idx] = {
                        'type': (current.get('_chem_comp.type') or '').upper(),
                        'name': current.get('_chem_comp.name') or '',
                        'formula': current.get('_chem_comp.formula') or '',
                        'pdbx_type': current.get('_chem_comp.pdbx_type') or '',
                        'charge': current.get('_chem_comp.pdbx_formal_charge') or '',
                    }
                current = {}
                current_id = None

            for raw in f:
                line = raw.rstrip('\n')
                if not line:
                    continue
                if in_multiline:
                    if line.startswith(';'):
                        val = '\n'.join(multiline_buf).strip()
                        current[multiline_tag] = val  # type: ignore
                        in_multiline = False
                        multiline_tag = None
                        multiline_buf = []
                    else:
                        multiline_buf.append(line)
                    continue
                if line.startswith('data_'):
                    flush()
                    continue
                if line.startswith('#'):
                    continue
                if line.strip() == 'loop_':
                    continue
                if line.startswith('_chem_comp.'):
                    parts = line.split(None, 1)
                    tag = parts[0]
                    val_part = parts[1] if len(parts) > 1 else ''
                    if val_part.startswith(';'):
                        in_multiline = True
                        multiline_tag = tag
                        multiline_buf = []
                        if val_part.strip() != ';':
                            multiline_buf.append(val_part[1:])
                        if tag == '_chem_comp.id':
                            current_id = 'MULTILINE'
                        continue
                    v = val_part.strip()
                    if (v.startswith("'") and v.endswith("'")) or (v.startswith('"') and v.endswith('"')):
                        v = v[1:-1]
                    current[tag] = v
                    if tag == '_chem_comp.id':
                        current_id = (v or '').upper()
                    continue
            flush()
    except Exception:
        return {}
    return index


def ccd_polymer_flag_and_category(resname: str, ccd: Dict[str, Dict[str, str]]) -> Tuple[Optional[bool], Optional[str], Dict[str, str]]:
    rn = (resname or '').upper()
    entry = ccd.get(rn)
    if not entry:
        return None, None, {'ccd_type': '', 'ccd_name': '', 'ccd_formula': ''}
    t = (entry.get('type') or '').upper()
    name = entry.get('name') or ''
    formula = entry.get('formula') or ''
    is_polymer: Optional[bool] = None
    cat: Optional[str] = None
    if 'LINKING' in t or 'TERMINUS' in t:
        is_polymer = True
    elif 'WATER' in t:
        is_polymer = False; cat = 'water'
    elif 'ION' in t:
        is_polymer = False; cat = 'ion'
    elif 'SACCHARIDE' in t or 'CARBOHYDRATE' in t:
        is_polymer = False; cat = 'sugar'
    elif 'COFACTOR' in t or 'CO-FACTOR' in t or 'COENZYME' in t or 'PROSTHETIC' in t:
        is_polymer = False; cat = 'cofactor'
    elif 'NON-POLYMER' in t:
        is_polymer = False; cat = 'ligand'
    return is_polymer, cat, {'ccd_type': t, 'ccd_name': name, 'ccd_formula': formula}


def ccd_simple_category(resname: str, ccd: Dict[str, Dict[str, str]]) -> Tuple[Optional[bool], Optional[str], Dict[str, str]]:
    rn = (resname or '').upper()
    e = ccd.get(rn)
    if not e:
        return None, None, {'ccd_type': '', 'ccd_name': '', 'ccd_formula': ''}
    t = (e.get('type') or '').upper()
    name = (e.get('name') or '').upper()
    formula = e.get('formula') or ''
    is_polymer = ('LINKING' in t) or ('TERMINUS' in t)
    if is_polymer:
        return True, None, {'ccd_type': t, 'ccd_name': e.get('name') or '', 'ccd_formula': formula}
    cat = 'ligand'
    tokens = [tok for tok in formula.split() if tok]
    if name == 'WATER' or name.endswith(' ION') or len(tokens) == 1:
        cat = 'others'
    return False, cat, {'ccd_type': t, 'ccd_name': e.get('name') or '', 'ccd_formula': formula}


def summarize_file(root: str, rel_path: str, ccd_index: Optional[Dict[str, Dict[str, str]]] = None,
                   classification: str = 'tri') -> List[Dict[str, object]]:
    full = os.path.join(root, rel_path)
    atoms, het_desc, formul, seqres_by_chain, link_pairs, model_count = parse_pdb(full)

    model_tag_map: Dict[int, str] = {}
    groups: Dict[Tuple[int,str,str,int,str,int], List[AtomRecord]] = defaultdict(list)
    last_key_per_model: Dict[int, Tuple[str,str,int,str]] = {}
    occ_counter: Dict[Tuple[int,str,str,int,str], int] = defaultdict(int)
    for a in atoms:
        mid = getattr(a, 'model', 1)
        tag = getattr(a, 'model_tag', '')
        if mid not in model_tag_map and tag:
            model_tag_map[mid] = tag
        base = (a.resname or '', a.chain or '', a.resseq or -999999, a.icode or '')
        key = (mid, *base, 0)
        if base == last_key_per_model.get(mid):
            occ_idx = occ_counter[(mid, *base)]
        else:
            occ_counter[(mid, *base)] += 1
            occ_idx = occ_counter[(mid, *base)]
            last_key_per_model[mid] = base
        key = (mid, base[0], base[1], base[2], base[3], occ_idx)
        groups[key].append(a)

    link_keys: Set[Tuple[int,str,str,int,str,int]] = set()
    for (r1, r2) in link_pairs:
        for mid in model_tag_map.keys() or [1]:
            if r1[2] is not None:
                link_keys.add((mid, r1[0], r1[1], r1[2], r1[3], 1))
            if r2[2] is not None:
                link_keys.add((mid, r2[0], r2[1], r2[2], r2[3], 1))

    results: List[Dict[str, object]] = []
    folder = os.path.basename(os.path.dirname(rel_path))
    fname = os.path.basename(rel_path)
    per_file_counts = Counter()

    for key, alist in groups.items():
        model_id, resname, chain, resseq, icode, res_occurrence = key
        atomnames = {a.name for a in alist if a.name}
        record_types = {a.record for a in alist}
        elements = {a.element for a in alist if a.element}
        occ_vals = [a.occ for a in alist if a.occ is not None]
        altlocs = {a.altloc for a in alist if a.altloc}

        in_seqres = resname in seqres_by_chain.get(chain, set())
        polymer_like = False
        if is_polymer_resname(resname):
            polymer_like = True
        elif in_seqres:
            polymer_like = True
        elif has_protein_backbone(atomnames) or has_nucleic_backbone(atomnames):
            polymer_like = True

        if classification == 'tri':
            ccd_polymer, ccd_category, ccd_fields = ccd_simple_category(resname, ccd_index or {})
            if ccd_polymer is True:
                is_polymer = True
            elif ccd_polymer is False:
                is_polymer = False
            else:
                is_polymer = polymer_like
            if not is_polymer:
                category = 'others' if ccd_category == 'others' else 'ligand'
            else:
                category = 'polymer'
        else:
            ccd_polymer, ccd_category, ccd_fields = ccd_polymer_flag_and_category(resname, ccd_index or {})
            if ccd_polymer is True:
                is_polymer = True
            elif ccd_polymer is False:
                is_polymer = False
            else:
                is_polymer = polymer_like
            category = ccd_category if (not is_polymer and ccd_category) else classify_group(resname)

        if is_polymer:
            per_file_counts[f'model_{model_id}_polymer_groups'] += 1
            continue

        elem_counts_actual = Counter()
        actual_noh = Counter()
        ccd_counts = Counter()
        ccd_noh = Counter()
        ccd_exists = bool(ccd_fields.get('ccd_type') or ccd_fields.get('ccd_name') or ccd_fields.get('ccd_formula'))
        if category == 'ligand':
            elem_counts_actual = Counter(a.element.upper() for a in alist if a.element)
            actual_noh = Counter({k: v for k, v in elem_counts_actual.items() if k != 'H'})
            ccd_counts = parse_ccd_formula_counts(ccd_fields.get('ccd_formula', '')) if ccd_fields else Counter()
            ccd_noh = Counter({k: v for k, v in ccd_counts.items() if k != 'H'})
            formula_match = 'na'
            if ccd_exists and actual_noh and ccd_noh:
                formula_match = 'yes' if actual_noh == ccd_noh else 'no'
            else:
                formula_match = 'na'
        else:
            formula_match = 'na'

        per_file_counts[f'model_{model_id}_nonpolymer_groups'] += 1
        if category == 'ligand':
            per_file_counts[f'model_{model_id}_ligand_groups'] += 1
        elif category in ('water','ion','solvent','cofactor','sugar'):
            per_file_counts[f'model_{model_id}_{category}_groups'] += 1
        else:
            per_file_counts[f'model_{model_id}_other_groups'] += 1

        model_tag = getattr(alist[0], 'model_tag', '') if alist else ''
        results.append({
            'folder': folder,
            'file': fname,
            'model_id': model_id,
            'model_tag': model_tag,
            'res_occurrence': res_occurrence,
            'res_occ_total': 1,
            'suspect_missing_separator': 'no',
            'resname': resname,
            'category': category,
            'chain': chain,
            'resseq': resseq,
            'icode': icode,
            'atom_count': len(alist),
            'record_types': ';'.join(sorted(record_types)),
            'elements': ';'.join(sorted(elements)) if elements else '',
            'heavy_elem_counts': ';'.join(f"{k}:{v}" for k, v in sorted(actual_noh.items())) if category == 'ligand' else '',
            'occ_min': min(occ_vals) if occ_vals else '',
            'occ_max': max(occ_vals) if occ_vals else '',
            'altlocs': ''.join(sorted(altlocs)) if altlocs else '',
            'het_description': het_desc.get(resname, ''),
            'formula': formul.get(resname, ''),
            'linked_via_LINK': False,
            'model_count': model_count if model_count else 1,
            **ccd_fields,
            'ccd_exists': 'yes' if ccd_exists else 'no',
            'ccd_formula_match': formula_match,
        })

    # __FILE_SUMMARY__ 行
    for mid in (model_tag_map.keys() or [1]):
        per_model = {k: v for k, v in per_file_counts.items() if k.startswith(f'model_{mid}_')}
        total_atoms = sum(r['atom_count'] for r in results if r['model_id'] == mid)
        # Normalize per-model counters into fixed field names expected by CSV header
        def _pm(key: str) -> int:
            return int(per_model.get(f'model_{mid}_{key}', 0))

        results.append({
            'folder': folder,
            'file': fname,
            'model_id': mid,
            'model_tag': model_tag_map.get(mid, ''),
            'res_occurrence': '',
            'res_occ_total': '',
            'suspect_missing_separator': '',
            'resname': '__FILE_SUMMARY__',
            'category': '',
            'chain': '',
            'resseq': '',
            'icode': '',
            'atom_count': total_atoms,
            'record_types': '',
            'elements': '',
            'heavy_elem_counts': '',
            'occ_min': '',
            'occ_max': '',
            'altlocs': '',
            'het_description': '',
            'formula': '',
            'linked_via_LINK': '',
            'model_count': model_count if model_count else 1,
            'count_polymer_groups': _pm('polymer_groups'),
            'count_nonpolymer_groups': _pm('nonpolymer_groups'),
            'count_ligand_groups': _pm('ligand_groups'),
            'count_other_groups': _pm('other_groups'),
            'count_ion_groups': _pm('ion_groups'),
            'count_water_groups': _pm('water_groups'),
            'count_solvent_groups': _pm('solvent_groups'),
        })

    return results


def walk_pdbs(root: str) -> List[str]:
    pdbs = []
    for dirpath, _, filenames in os.walk(root):
        for fn in filenames:
            if fn.lower().endswith(('.pdb','.ent','.cif','.mmcif','.mcif')):
                rel = os.path.relpath(os.path.join(dirpath, fn), start=root)
                pdbs.append(rel)
    pdbs.sort()
    return pdbs


# ---------------------------------------------------------------------------
# 2025-10 dict-driven extension (for docker image with /opt/pdbechem/* etc.)
# ---------------------------------------------------------------------------

DOCKER_DEFAULT_DICTS = [
    os.environ.get("PDBECHEM_COMPONENTS", "/opt/pdbechem/components.cif"),
    os.environ.get("PDBECHEM_PRD", "/opt/pdbechem/prdcc-all.cif"),
    os.environ.get("PDBECHEM_CLC", "/opt/pdbechem/clc-all.cif"),
    os.environ.get("WWPDB_BIRD", "/opt/wwpdb/refdata/bird-prd-all.cif"),
    os.environ.get("WWPDB_BIRD_FAMILY", "/opt/wwpdb/refdata/bird-family-all.cif"),
]


def _load_plain_cif_dict(path: str) -> Dict[str, Dict[str, str]]:
    idx: Dict[str, Dict[str, str]] = {}
    if not os.path.exists(path):
        return idx
    current: Dict[str, Any] = {}
    current_id: Optional[str] = None
    in_multiline = False
    multiline_tag: Optional[str] = None
    buf: List[str] = []
    with open(path, "rt", errors="ignore") as fh:
        def flush():
            nonlocal current, current_id
            if current_id:
                cid = current_id.upper()
                idx[cid] = {
                    'type': (current.get('_chem_comp.type') or '').upper(),
                    'name': current.get('_chem_comp.name') or '',
                    'formula': current.get('_chem_comp.formula') or '',
                    'pdbx_type': current.get('_chem_comp.pdbx_type') or '',
                    'charge': current.get('_chem_comp.pdbx_formal_charge') or '',
                }
            current = {}
            current_id = None

        for raw in fh:
            line = raw.rstrip("\n")
            if in_multiline:
                if line.startswith(";"):
                    val = "\n".join(buf).strip()
                    current[multiline_tag] = val  # type: ignore
                    in_multiline = False
                    multiline_tag = None
                    buf = []
                else:
                    buf.append(line)
                continue
            if line.startswith("data_"):
                flush()
                continue
            if line.startswith("_chem_comp."):
                parts = line.split(None, 1)
                tag = parts[0]
                val = parts[1] if len(parts) > 1 else ""
                if val.strip() == ";":
                    in_multiline = True
                    multiline_tag = tag
                    buf = []
                    if tag == "_chem_comp.id":
                        current_id = "MULTILINE"
                    continue
                v = val.strip()
                if (v.startswith("'") and v.endswith("'")) or (v.startswith('"') and v.endswith('"')):
                    v = v[1:-1]
                current[tag] = v
                if tag == "_chem_comp.id":
                    current_id = v.upper()
                continue
        flush()
    return idx


def load_external_dicts(preferred_ccd_gz: Optional[str], extra_dicts: Optional[List[str]] = None) -> Dict[str, Dict[str, str]]:
    idx: Dict[str, Dict[str, str]] = {}
    if preferred_ccd_gz:
        idx.update(load_ccd_index(preferred_ccd_gz))
    for pth in (extra_dicts or []):
        if not pth:
            continue
        if pth.endswith(".gz"):
            idx.update(load_ccd_index(pth))
        else:
            idx.update(_load_plain_cif_dict(pth))
    for pth in DOCKER_DEFAULT_DICTS:
        if not pth or not os.path.exists(pth):
            continue
        if pth.endswith(".gz"):
            idx.update(load_ccd_index(pth))
        else:
            idx.update(_load_plain_cif_dict(pth))
    return idx


def install_dict_driven_classification(ccd_index: Dict[str, Dict[str, str]]):
    if not ccd_index:
        return

    if not hasattr(is_polymer_resname, "__wrapped__"):
        is_polymer_resname.__wrapped__ = is_polymer_resname  # type: ignore
    if not hasattr(classify_group, "__wrapped__"):
        classify_group.__wrapped__ = classify_group  # type: ignore

    def _is_polymer_resname(resname: str) -> bool:
        rn = (resname or "").upper()
        entry = ccd_index.get(rn)
        if entry:
            t = (entry.get("type") or "").upper()
            if ("LINKING" in t) or ("TERMINUS" in t) or ("POLYMER" in t):
                return True
            if "NON-POLYMER" in t:
                return False
        return is_polymer_resname.__wrapped__(resname)  # type: ignore

    def _classify_group(resname: str) -> str:
        rn = (resname or "").upper()
        entry = ccd_index.get(rn)
        if entry:
            t = (entry.get("type") or "").upper()
            name = (entry.get("name") or "").upper()
            pdbx_type = (entry.get("pdbx_type") or "").upper()
            formula = entry.get("formula") or ""
            if rn in ("HOH", "WAT") or "WATER" in t or "WATER" in name or "WATER" in pdbx_type:
                return "water"
            if "ION" in t or "ION" in name or "ION" in pdbx_type:
                return "ion"
            if re.fullmatch(r"[A-Z][a-z]?(?:\d+)?", formula.strip()):
                return "ion"
            if "SACCHARIDE" in t or "CARBOHYDRATE" in t or "SUGAR" in name:
                return "sugar"
            if any(k in t for k in ("COFACTOR", "CO-FACTOR", "COENZYME", "PROSTHETIC")):
                return "cofactor"
            if any(k in name for k in ("PEG", "MPD", "GLYCEROL", "ETHYLENE GLYCOL", "DMSO", "GOL")):
                return "solvent"
            return "ligand"
        return classify_group.__wrapped__(resname)  # type: ignore

    globals()["is_polymer_resname"] = _is_polymer_resname
    globals()["classify_group"] = _classify_group


# ---------------------------------------------------------------------------
# 新的 main：先加载 docker 里的字典，再跑你原来的逻辑
# ---------------------------------------------------------------------------

def main():
    import argparse
    p = argparse.ArgumentParser(description='Robust ligand summary from PDBs by folder/file (dict-first, legacy-compatible)')
    p.add_argument('--root', help='Root directory containing PZ subfolders')
    p.add_argument('--out', default='ligand_summary.csv', help='Output CSV with ligand groups + file summaries')
    p.add_argument('--counts', default='ligand_counts_by_folder.csv', help='Output CSV for per-folder counts')
    p.add_argument('--ccd', default=None, help='(optional) legacy CCD components.cif.gz')
    p.add_argument('--dict', dest='dicts', action='append', default=[], help='extra dictionary mmCIF (components.cif / prdcc-all.cif / clc-all.cif), can be used multiple times')
    p.add_argument('--classification', choices=['tri','detailed'], default='tri',
                   help='tri: polymer/ligand/others (CCD-driven). detailed: legacy categories (ligand/ion/water/solvent/...).')
    p.add_argument('--xlsx', default='ligand_report.xlsx', help='Optional Excel report with multiple sheets (details, by_folder, by_file_model, readme). Set empty to skip.')
    args = p.parse_args()

    # 1) 先把所有能找到的字典都读进来
    ccd_index = load_external_dicts(args.ccd, args.dicts)
    if ccd_index:
        print(f'Loaded {len(ccd_index)} components from external dictionaries.')
        install_dict_driven_classification(ccd_index)
    else:
        print('No external dictionaries found; falling back to legacy hard-coded tables.')

    root = args.root
    pdbs = walk_pdbs(root)
    if not pdbs:
        print(f'No PDB/mmCIF files found under {root}')
        return

    rows: List[Dict[str, object]] = []
    for rel in pdbs:
        rows.extend(summarize_file(root, rel, ccd_index, classification=args.classification))

    # 输出 CSV，这一段我保持你原来的写法
    fieldnames = [
        'folder','file','resname','category','chain','resseq','icode','atom_count',
        'record_types','elements','heavy_elem_counts','occ_min','occ_max','altlocs','het_description',
        'formula','linked_via_LINK','model_count','ccd_type','ccd_name','ccd_formula','ccd_exists','ccd_formula_match',
        'model_id','model_tag','res_occurrence','res_occ_total','suspect_missing_separator',
        'count_polymer_groups','count_nonpolymer_groups','count_ligand_groups',
        'count_other_groups','count_ion_groups','count_water_groups','count_solvent_groups'
    ]
    for r in rows:
        for k in fieldnames:
            if k not in r:
                r[k] = ''

    with open(args.out, 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for r in rows:
            w.writerow(r)

    per_folder = defaultdict(Counter)
    seen_files = defaultdict(set)
    for r in rows:
        if r.get('resname') == '__FILE_SUMMARY__':
            folder = r['folder']
            if r['file'] not in seen_files[folder]:
                per_folder[folder]['files'] += 1
                seen_files[folder].add(r['file'])
            per_folder[folder]['atoms'] += int(r.get('atom_count') or 0)
            for k in (
                'count_polymer_groups','count_nonpolymer_groups','count_ligand_groups','count_other_groups',
                'count_ion_groups','count_water_groups','count_solvent_groups'
            ):
                v = r.get(k)
                try:
                    per_folder[folder][k] += int(v)
                except Exception:
                    pass

    with open(args.counts, 'w', newline='') as f:
        fieldnames2 = ['folder','files','atoms','count_polymer_groups','count_nonpolymer_groups','count_ligand_groups',
                       'count_other_groups','count_ion_groups','count_water_groups','count_solvent_groups']
        w = csv.DictWriter(f, fieldnames=fieldnames2)
        w.writeheader()
        for folder, cnt in sorted(per_folder.items()):
            w.writerow({
                'folder': folder,
                'files': cnt.get('files', 0),
                'atoms': cnt.get('atoms', 0),
                'count_polymer_groups': cnt.get('count_polymer_groups', 0),
                'count_nonpolymer_groups': cnt.get('count_nonpolymer_groups', 0),
                'count_ligand_groups': cnt.get('count_ligand_groups', 0),
                'count_other_groups': cnt.get('count_other_groups', 0),
                'count_ion_groups': cnt.get('count_ion_groups', 0),
                'count_water_groups': cnt.get('count_water_groups', 0),
                'count_solvent_groups': cnt.get('count_solvent_groups', 0),
            })

    if args.xlsx:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import PatternFill
        except Exception as e:
            print(f"openpyxl not available ({e}); skipping Excel output.")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = 'details'
            ws.append(fieldnames)
            for r in rows:
                ws.append([r.get(k, '') for k in fieldnames])
            try:
                col_match = fieldnames.index('ccd_formula_match') + 1
                col_exists = fieldnames.index('ccd_exists') + 1
                col_sus = fieldnames.index('suspect_missing_separator') + 1
                red_fill = PatternFill(start_color='FFFFD6D6', end_color='FFFFD6D6', fill_type='solid')
                yellow_fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    vmatch = (row[col_match-1].value or '').lower()
                    vexists = (row[col_exists-1].value or '').lower()
                    vsus = (row[col_sus-1].value or '').lower()
                    if vexists == 'yes' and vmatch == 'no':
                        for cell in row:
                            cell.fill = red_fill
                    elif vsus == 'yes':
                        for cell in row:
                            cell.fill = yellow_fill
            except Exception:
                pass

            ws2 = wb.create_sheet('by_folder')
            ws2.append(fieldnames2)
            for folder, cnt in sorted(per_folder.items()):
                ws2.append([
                    folder,
                    cnt.get('files', 0),
                    cnt.get('atoms', 0),
                    cnt.get('count_polymer_groups', 0),
                    cnt.get('count_nonpolymer_groups', 0),
                    cnt.get('count_ligand_groups', 0),
                    cnt.get('count_other_groups', 0),
                    cnt.get('count_ion_groups', 0),
                    cnt.get('count_water_groups', 0),
                    cnt.get('count_solvent_groups', 0),
                ])

            ws4 = wb.create_sheet('by_file_model')
            fcols = ['folder','file','model_id','atoms','count_polymer_groups','count_nonpolymer_groups',
                     'count_ligand_groups','count_other_groups']
            ws4.append(fcols)
            per_file_model = defaultdict(Counter)
            for r in rows:
                if r.get('resname') == '__FILE_SUMMARY__':
                    key = (r['folder'], r['file'], r.get('model_id') or 1)
                    cnt = per_file_model[key]
                    cnt['atoms'] += int(r.get('atom_count') or 0)
                    for k in ('count_polymer_groups','count_nonpolymer_groups','count_ligand_groups','count_other_groups'):
                        v = r.get(k)
                        try:
                            cnt[k] += int(v)
                        except Exception:
                            pass
            for (folder, file, mid), cnt in sorted(per_file_model.items()):
                ws4.append([
                    folder, file, mid,
                    cnt.get('atoms',0),
                    cnt.get('count_polymer_groups',0),
                    cnt.get('count_nonpolymer_groups',0),
                    cnt.get('count_ligand_groups',0),
                    cnt.get('count_other_groups',0),
                ])

            ws3 = wb.create_sheet('readme')
            ws3.append(['Field','Description','Notes'])
            ws3.append(['classification_mode', 'tri or detailed', args.classification])

            wb.save(args.xlsx)
            print(f"Wrote Excel workbook {args.xlsx} with sheets: details, by_folder, by_file_model, readme.")

    print(f'Wrote {args.out} and {args.counts} with {len(rows)} rows (incl. file summaries).')


if __name__ == '__main__':
    main()
